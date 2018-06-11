'''
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def createHeaderRow(sheet, titleHeaders, headerCells, fillColor, textRotation):
    headerFont = Font(bold = True)
    alignment = Alignment(text_rotation = textRotation)
    cellFill = PatternFill(fill_type = 'solid', start_color = fillColor, end_color = fillColor)
    for i in range(0, len(titleHeaders)):
        sheet[headerCells[i]] = titleHeaders[i]
        sheet[headerCells[i]].font = headerFont
        sheet[headerCells[i]].alignment = alignment
        sheet[headerCells[i]].fill = cellFill

    wb.save('My_book_with_Yellow_Tab.xlsx')

wb = Workbook()
ws = wb.create_sheet('My_Color_Title',0)
ws.sheet_properties.tabColor = 'FFFF00'
headers=["JavaVersion", "Spring Used?", "Hibernate Used?"]
headerCells = ["A1","B1","C1"]
createHeaderRow(wb['My_Color_Title'], headers, headerCells, '3D9407', 45)

ws = wb.create_sheet('New Sheet',1)
ws.sheet_properties.tabColor = 'FF0000'
headers=["MNGT_NAME", "DotNetVersion"]
headerCells = ["A1","B1"]
cellsRange = ws["A1","B1"]

createHeaderRow(wb['New Sheet'], headers, headerCells, 'E2EC0C', 45)

'''
from binaryornot.check import is_binary
if is_binary(r"C:\Users\VRE\Documents\My Received Files\AIO Journal DM.xdmz"):
    print("True")
else:
    print("False")

